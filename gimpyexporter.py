#!/usr/bin/python



# Written by Alex Hernandez
# 09/17/2015
""" This script will read an excel file and will automatically
    upload the test cases to GIMPY, Update the feature suite, and update your excel file.
"""
# Import all necessary modules.
import xlrd
from collections import OrderedDict
import simplejson as json
import shutil
from openpyxl import load_workbook
import bll
import sys
# Title
print "\n"
print "                                             |---- GIMPY EXPORTER v1.2 ----|"
print "\n"
print "\n"

"""Login to GIMPY"""

print "   -Login into GIMPY-"

print "\n"

# Stores the username in username variable
username = raw_input(" > What is your username: ")

# Logs user in and stores the token
loginToken = bll.db.login_gimpy(username)

print "\n"

print "  -Login Successful-"

print "\n"

"""Check Paths"""

# Check path config
bll.db.checkPathConfig()

# Set directory path
dirPath = bll.db.setDirPath()

# Set template path
templatePath = bll.db.setTemplatePath()


# Set excel path
excelPath = bll.db.setExcelPath()

print "\n"

# Set Test Case set

name = raw_input("What do you want to name this set of Test Cases?: ")

# Set Test Case Set name

testCaseSet = dirPath + name
print "\n"
print "Your test cases will be stored here: " + testCaseSet

# Create testCaseSet

bll.db.setTSname(testCaseSet)

print "\n"
print "\n"
print "\n"

# Declare the starting position of the test cases that have been created.
ExcelQuestion = raw_input("*** Please save and close the Excel file *** \n\n"
                          "   Continue?: (Y/N)" ).lower()
if ExcelQuestion == "y":
    print "\n"
    # Declare the starting position of the test cases that have been created.
    start_row = int(raw_input(" > What row does your first test case start at: ")) - 1
    while start_row <= 1:
        start_row = int(raw_input("     > *** Please enter a number higher than 1 or 2: *** ")) - 1
else:
    sys.exit(0)

print "\n"

print "\n"

print "---------------------------- TEST CASE RESULTS---------------------------------"

print "\n"

"""Open Excel File"""

# Open the workbook and select the first worksheet
wb = bll.db.openWB(excelPath)
sh = bll.db.openWS(wb)

# List to hold dictionaries
test_cases_list = []

# List to hold feature suite id's
fs_list = []

# List to hold test Case id's
test_case_id = []

""" Iterate through each row in worksheet and fetch values into dict  """

for rownum in range(start_row, sh.nrows):
    test_case = OrderedDict()
    row_values = sh.row_values(rownum)
    test_case['Name'] = row_values[0]
    test_case['Description'] = row_values[1]
    test_case['Assumption'] = row_values[2]
    test_case['Limitation'] = row_values[3]
    test_case['Note'] = row_values[4]
    test_case['EstimatedTimeToRun'] = row_values[5]
    test_case['Data'] = row_values[6]
    test_case['State ID'] = int(row_values[7])
    test_case['Product ID'] = int(row_values[8])
    test_case['Priority ID'] = int(row_values[9])
    test_case['TestType ID'] = int(row_values[10])
    test_case['HowFound ID'] = int(row_values[11])
    test_case['CodeBase ID'] = int(row_values[12])
    test_case['Release ID'] = int(row_values[13])
    test_case['Action'] = row_values[14]
    test_case['Result'] = row_values[15]
    test_case['Comment'] = row_values[16]


    # Append the values to the test_cases_list
    test_cases_list.append(test_case)

    # Append the values to the fs_list
    fs_list.append(int(row_values[17]))

"""
    Iterate through every item on the list, create a file for it
    and copy the information from the excel information to the template file copied.
"""

for i in range(len(test_cases_list)):
    shutil.copyfile(templatePath, testCaseSet + '/%s%d.json' % (name,i))
    # Opens, loads file we want to be written and stores it in the "data" variable
    jsonFile = open(testCaseSet + '/%s%d.json' % (name,i), "r")
    data = json.load(jsonFile, object_pairs_hook=OrderedDict)
    jsonFile.close()

    # Fetches each entity in JSON file and replaces element with new imported information
    data["Name"] = test_cases_list[i]['Name']
    data['Description'] = test_cases_list[i]['Description']
    data['Assumption'] = test_cases_list[i]['Assumption']
    data['Limitation'] =  test_cases_list[i]['Limitation']
    data['Note'] = test_cases_list[i]['Note']
    data['EstimatedTimeToRun'] = test_cases_list[i]['EstimatedTimeToRun']
    data['Data'] = test_cases_list[i]['Data']
    data['State']['Id'] = test_cases_list[i]['State ID']
    data['Product']['Id'] = test_cases_list[i]['Product ID']
    data['Priority']['Id'] = test_cases_list[i]['Priority ID']
    data['TestType']['Id'] = test_cases_list[i]['TestType ID']
    data['HowFound']['Id'] = test_cases_list[i]['HowFound ID']
    data['CodeBase']['Id'] = test_cases_list[i]['CodeBase ID']
    data['Release']['Id'] = test_cases_list[i]['Release ID']
    data['Steps'][0]['Action'] = test_cases_list[i]['Action']
    data['Steps'][0]['ExpectedResult'] = test_cases_list[i]['Result']
    data['Steps'][0]['Comment'] = test_cases_list[i]['Comment']


    """Create the Test Case"""

    addtestcase_response = bll.db.add_a_test_case(data, loginToken)


    """Update the Feature suite"""

    # fs_id will hold the value of the id in the fs_list
    fs_id = str(fs_list[i])

    updatefs_response = bll.db.update_feature_suite(fs_id, addtestcase_response['Id'], loginToken)


    # Print the results of test cases of each test case that has been created
    print "Test Case"
    print "---------"
    print "" +    str(data['Name'])
    print "\n"
    print "Status Code: " + str(updatefs_response)
    print "Test Case ID: " + str(addtestcase_response['Id'])
    print "Feature Suite " + fs_id + ": Updated successfully with TC" + str(addtestcase_response['Id'])


    # Append the responses to test_case_id list to write to excel file
    test_case_id.append(addtestcase_response['Id'])


    """Writes new information to new file and closes it"""

    jsonFile = open(testCaseSet + "/%s%d.json" % (name,i), "w+")
    jsonFile.write(json.dumps(data['Name'], indent=4 * ' '))
    jsonFile.write(json.dumps(addtestcase_response, indent=4 * ' '))
    jsonFile.close()

    print "\n"

"""Writes the ID back to the excel file"""
# Load the excel file
wb = load_workbook(filename=excelPath)
# Get the sheet
ws = wb.get_sheet_by_name("Sheet1")
row1 = start_row + 1
for item in range(len(test_case_id)):
    # Write the test case id to the excel file
    ws['S'+str(row1)] = test_case_id[item]
    row1 = row1 + 1

# save the excel file
wb.save(excelPath)

print "\n"

print "---------------------------- EXPORTS COMPLETED---------------------------------"

print "\n"















